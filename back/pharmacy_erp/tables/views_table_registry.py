# API Views for Table Registry

from rest_framework import generics, filters, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q
from django.apps import apps
from django.db import models

from tables.model_definitions.table_registry import TableRegistry


SYSTEM_FIELD_NAMES = {
    'id',
    'pk',
    'created_at',
    'updated_at',
    'created_by',
    'updated_by',
}


def _field_type_to_schema(field):
    """Map Django model fields to the registry column type values."""
    if isinstance(field, (models.AutoField, models.BigAutoField, models.SmallAutoField)):
        return 'integer'
    if isinstance(field, (models.IntegerField, models.PositiveIntegerField, models.PositiveSmallIntegerField, models.SmallIntegerField)):
        return 'integer'
    if isinstance(field, (models.DecimalField, models.FloatField)):
        return 'decimal'
    if isinstance(field, models.BooleanField):
        return 'boolean'
    if isinstance(field, models.DateTimeField):
        return 'datetime'
    if isinstance(field, models.DateField):
        return 'date'
    if isinstance(field, models.EmailField):
        return 'email'
    if isinstance(field, models.TextField):
        return 'text'
    if isinstance(field, (models.ForeignKey, models.OneToOneField)):
        return 'reference'
    return 'string'


def _is_required_field(field):
    if getattr(field, 'primary_key', False):
        return False
    if not getattr(field, 'editable', True):
        return False
    if getattr(field, 'auto_now', False) or getattr(field, 'auto_now_add', False):
        return False
    if getattr(field, 'null', False) or getattr(field, 'blank', False):
        return False
    if getattr(field, 'has_default', lambda: False)():
        return False
    return True


def _can_import_to_field(field):
    if getattr(field, 'primary_key', False):
        return False
    if not getattr(field, 'editable', True):
        return False
    if getattr(field, 'auto_now', False) or getattr(field, 'auto_now_add', False):
        return False
    if field.name in SYSTEM_FIELD_NAMES:
        return False
    return True


def _build_column_from_model_field(field):
    raw_choices = getattr(field, 'choices', None) or []
    choices = [choice_value for choice_value, _ in raw_choices] or None
    return {
        'name': field.name,
        'label': getattr(field, 'verbose_name', field.name).replace('_', ' ').title(),
        'type': _field_type_to_schema(field),
        'required': _is_required_field(field),
        'description': getattr(field, 'help_text', '') or '',
        'example': '',
        'include_in_import': _can_import_to_field(field),
        'include_in_export': True,
        'is_identifier': bool(getattr(field, 'unique', False) or field.name in {'name', 'code', 'barcode', 'naming_series'}),
        'choices': choices,
        'max_length': getattr(field, 'max_length', None),
        'null': getattr(field, 'null', False),
        'blank': getattr(field, 'blank', False),
        'model_field': True,
    }


def _get_model_fields_for_table(table):
    if not table.model_class or '.' not in table.model_class:
        return []

    try:
        app_label, model_name = table.model_class.split('.', 1)
        model = apps.get_model(app_label, model_name)
    except (LookupError, ValueError):
        return []

    fields = []
    for field in model._meta.get_fields():
        if getattr(field, 'auto_created', False) and not getattr(field, 'concrete', False):
            continue
        if not getattr(field, 'concrete', False):
            continue
        fields.append(field)
    return fields


def get_resolved_table_columns(table):
    """
    Merge saved registry column config with actual Django model fields.
    Saved settings win, but missing model fields are auto-added so the UI
    can manage importability for every real column in the table.
    """
    saved_columns = table.columns or []
    saved_by_name = {column.get('name'): column for column in saved_columns if column.get('name')}

    resolved_columns = []
    model_fields = _get_model_fields_for_table(table)

    for field in model_fields:
        model_column = _build_column_from_model_field(field)
        saved_column = saved_by_name.pop(field.name, {})
        resolved_columns.append({
            **model_column,
            **saved_column,
            'name': field.name,
            'label': saved_column.get('label') or model_column['label'],
            'type': saved_column.get('type') or model_column['type'],
            'choices': saved_column.get('choices', model_column['choices']),
            'max_length': saved_column.get('max_length', model_column['max_length']),
        })

    for column in saved_columns:
        column_name = column.get('name')
        if column_name and column_name in saved_by_name:
            resolved_columns.append({
                'include_in_import': True,
                'include_in_export': True,
                'is_identifier': False,
                'required': False,
                'description': '',
                'example': '',
                'choices': None,
                'max_length': None,
                'model_field': False,
                **column,
            })

    return resolved_columns


class TableRegistryListView(generics.ListAPIView):
    """List all table registry entries with optional filtering."""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = TableRegistry.objects.filter(is_active=True)

        # Filter by module
        module = self.request.query_params.get('module', None)
        if module:
            queryset = queryset.filter(module=module)

        # Filter by submodule
        submodule = self.request.query_params.get('submodule', None)
        if submodule:
            queryset = queryset.filter(submodule=submodule)

        # Filter by importable
        importable = self.request.query_params.get('importable', None)
        if importable is not None:
            queryset = queryset.filter(importable=importable.lower() == 'true')

        # Filter by exportable
        exportable = self.request.query_params.get('exportable', None)
        if exportable is not None:
            queryset = queryset.filter(exportable=exportable.lower() == 'true')

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Serialize manually for more control
        data = []
        for table in queryset:
            resolved_columns = get_resolved_table_columns(table)
            data.append({
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'module': table.module,
                'submodule': table.submodule,
                'frontend_path': table.frontend_path,
                'backend_endpoint': table.backend_endpoint,
                'keywords': table.get_keywords_list(),
                'importable': table.importable,
                'exportable': table.exportable,
                'description': table.description,
                'icon': table.icon,
                'db_table_name': table.db_table_name,
                'model_class': table.model_class,
                'columns': resolved_columns,
            })

        return Response(data)


class TableRegistrySearchView(APIView):
    """Search tables by query string."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        query = request.query_params.get('q', '').strip()
        module = request.query_params.get('module', None)

        if not query:
            return Response([])

        queryset = TableRegistry.objects.filter(is_active=True)

        if module:
            queryset = queryset.filter(module=module)

        # Search in multiple fields
        queryset = queryset.filter(
            Q(table_name__icontains=query) |
            Q(table_code__icontains=query) |
            Q(description__icontains=query) |
            Q(search_keywords__icontains=query)
        ).distinct()

        # Serialize results
        data = []
        for table in queryset:
            # Calculate relevance score (exact matches score higher)
            score = 0
            q_lower = query.lower()
            if q_lower == table.table_name.lower():
                score = 100
            elif q_lower == table.table_code.lower():
                score = 90
            elif q_lower in table.table_name.lower():
                score = 70
            elif q_lower in table.table_code.lower():
                score = 60
            else:
                score = 50

            data.append({
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'module': table.module,
                'submodule': table.submodule,
                'frontend_path': table.frontend_path,
                'keywords': table.get_keywords_list(),
                'description': table.description,
                'icon': table.icon,
                'relevance_score': score,
            })

        # Sort by relevance score
        data.sort(key=lambda x: x['relevance_score'], reverse=True)

        return Response(data)


class TableRegistryDetailView(APIView):
    """Get details of a specific table by code."""
    permission_classes = [IsAuthenticated]

    def get(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code, is_active=True)
            resolved_columns = get_resolved_table_columns(table)
            data = {
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'module': table.module,
                'submodule': table.submodule,
                'frontend_path': table.frontend_path,
                'backend_endpoint': table.backend_endpoint,
                'keywords': table.get_keywords_list(),
                'importable': table.importable,
                'exportable': table.exportable,
                'import_template': table.import_template,
                'description': table.description,
                'icon': table.icon,
                'db_table_name': table.db_table_name,
                'model_class': table.model_class,
                'requires_auth': table.requires_auth,
                'required_permissions': table.required_permissions,
                'columns': resolved_columns,
            }
            return Response(data)
        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class TableRegistryModulesView(APIView):
    """Get list of all modules."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        modules = TableRegistry.objects.filter(
            is_active=True
        ).values_list('module', flat=True).distinct()

        return Response(sorted(modules))


class TableRegistryCreateView(APIView):
    """Create a new table registry entry."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        data = request.data
        try:
            table = TableRegistry.objects.create(**data)
            return Response({
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'message': 'Table created successfully'
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


class TableRegistryUpdateView(APIView):
    """Update an existing table registry entry."""
    permission_classes = [IsAuthenticated]

    def put(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code)
            data = request.data

            if 'columns' in data and isinstance(data['columns'], list):
                resolved_columns = get_resolved_table_columns(table)
                allowed_column_names = {
                    column.get('name') for column in resolved_columns if column.get('name')
                }
                data['columns'] = [
                    column for column in data['columns']
                    if column.get('name') and column.get('name') in allowed_column_names
                ]

            # Update fields
            for key, value in data.items():
                if hasattr(table, key):
                    setattr(table, key, value)

            table.save()
            return Response({
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'message': 'Table updated successfully'
            })
        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class TableRegistryDeleteView(APIView):
    """Delete a table registry entry."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code)
            if table.is_system:
                return Response(
                    {'error': 'Cannot delete system tables'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            table.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class ImportableTablesView(APIView):
    """Get list of tables that support import."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tables = TableRegistry.objects.filter(
            is_active=True,
            importable=True
        )

        data = []
        for table in tables:
            data.append({
                'id': table.table_code,
                'table_code': table.table_code,
                'table_name': table.table_name,
                'module': table.module,
                'import_template': table.import_template,
                'model_class': table.model_class,
                'backend_endpoint': table.backend_endpoint,
            })

        return Response(data)


class TableSchemaView(APIView):
    """Get schema information for a table (fields, types, required fields)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code, is_active=True)

            if not table.importable:
                return Response(
                    {'error': 'Table does not support import'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Use columns from registry (all columns for schema display)
            fields = get_resolved_table_columns(table)

            # If no columns defined, fall back to generic schema
            if not fields:
                fields = [
                    {
                        'name': 'name',
                        'label': 'Name',
                        'type': 'string',
                        'required': True,
                        'example': 'Sample Name',
                        'is_identifier': True,
                        'include_in_import': True,
                    },
                    {
                        'name': 'description',
                        'label': 'Description',
                        'type': 'string',
                        'required': False,
                        'example': 'Sample description',
                        'include_in_import': True,
                    },
                ]

            # Filter to only include fields marked for import in the schema
            import_fields = [f for f in fields if f.get('include_in_import', True)]

            return Response({
                'table_code': table_code,
                'table_name': table.table_name,
                'fields': import_fields,
                'required_fields': [f['name'] for f in import_fields if f.get('required')],
                'identifier_fields': [f['name'] for f in import_fields if f.get('is_identifier')],
                'all_columns': fields,  # Include all columns for reference
            })

        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )

class TableImportPreviewView(APIView):
    """Preview import data before importing."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code, is_active=True)

            if not table.importable:
                return Response(
                    {'error': 'Table does not support import'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded_file = request.FILES.get('file')
            mode = request.data.get('mode', 'insert')

            if not uploaded_file:
                return Response(
                    {'error': 'No file provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Parse CSV or Excel
            import csv
            import io

            valid_rows = []
            invalid_rows = []
            total_count = 0

            # Get import columns from registry
            import_columns = [
                c for c in get_resolved_table_columns(table)
                if c.get('include_in_import', True)
            ]
            required_columns = [c['name'] for c in import_columns if c.get('required')]

            try:
                file_name = uploaded_file.name.lower()

                if file_name.endswith('.csv'):
                    # Handle CSV
                    content = uploaded_file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = list(reader)
                elif file_name.endswith(('.xlsx', '.xls')):
                    # Handle Excel - requires openpyxl
                    try:
                        import openpyxl
                        uploaded_file.seek(0)
                        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                        ws = wb.active

                        # Get headers from first row
                        headers = [cell.value for cell in ws[1]]

                        # Convert to dict rows
                        rows = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            row_dict = {}
                            for i, header in enumerate(headers):
                                if header:  # Skip empty headers
                                    row_dict[header] = row[i] if i < len(row) else None
                            rows.append(row_dict)
                    except ImportError:
                        return Response(
                            {'error': 'Excel support requires openpyxl. Please install it with: pip install openpyxl'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    return Response(
                        {'error': 'Unsupported file format. Please upload CSV or Excel (.xlsx, .xls) files.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                for idx, row in enumerate(rows, start=2):
                    total_count += 1
                    errors = []

                    # Convert row values to strings for validation
                    row = {k: str(v) if v is not None else '' for k, v in row.items()}

                    # Validate required columns
                    for col_name in required_columns:
                        if not row.get(col_name) or not str(row.get(col_name)).strip():
                            errors.append(f"{col_name} is required")

                    # Validate data types
                    for col in import_columns:
                        col_name = col['name']
                        value = row.get(col_name)
                        if value:
                            col_type = col.get('type', 'string')
                            if col_type == 'integer':
                                try:
                                    int(value)
                                except ValueError:
                                    errors.append(f"{col_name} must be a number")
                            elif col_type == 'decimal':
                                try:
                                    float(value)
                                except ValueError:
                                    errors.append(f"{col_name} must be a decimal number")
                            elif col_type == 'choice':
                                choices = col.get('choices', [])
                                if choices and value not in choices:
                                    errors.append(f"{col_name} must be one of: {', '.join(choices)}")

                    if errors:
                        invalid_rows.append({
                            'row': idx,
                            'data': row,
                            'errors': errors
                        })
                    else:
                        valid_rows.append(row)

            except Exception as e:
                return Response(
                    {'error': f'Failed to parse file: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            return Response({
                'valid_rows': valid_rows,
                'invalid_rows': invalid_rows,
                'total_count': total_count,
                'mode': mode,
            })

        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class TableImportExecuteView(APIView):
    """Execute the actual import."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request, table_code):
        try:
            table = TableRegistry.objects.get(table_code=table_code, is_active=True)

            if not table.importable:
                return Response(
                    {'error': 'Table does not support import'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            uploaded_file = request.FILES.get('file')
            mode = request.data.get('mode', 'insert')

            if not uploaded_file:
                return Response(
                    {'error': 'No file provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Parse and import
            import csv
            import io

            created = 0
            updated = 0
            errors = 0

            try:
                content = uploaded_file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))

                # Try to get model
                model = None
                if table.model_class:
                    try:
                        model = apps.get_model(*table.model_class.split('.'))
                    except:
                        pass

                for row in reader:
                    try:
                        if model:
                            if mode == 'insert':
                                model.objects.create(**row)
                                created += 1
                            elif mode == 'update':
                                identifier = row.get('naming_series') or row.get('id') or row.get('name')
                                if identifier:
                                    existing = model.objects.filter(naming_series=identifier).first()
                                    if existing:
                                        for key, value in row.items():
                                            setattr(existing, key, value)
                                        existing.save()
                                        updated += 1
                                    else:
                                        created += 1
                                        model.objects.create(**row)
                                else:
                                    created += 1
                                    model.objects.create(**row)
                        else:
                            created += 1
                    except Exception as e:
                        errors += 1
                        print(f"Import error: {e}")

            except Exception as e:
                return Response(
                    {'error': f'Import failed: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            return Response({
                'created': created,
                'updated': updated,
                'errors': errors,
            })

        except TableRegistry.DoesNotExist:
            return Response(
                {'error': 'Table not found'},
                status=status.HTTP_404_NOT_FOUND
            )
